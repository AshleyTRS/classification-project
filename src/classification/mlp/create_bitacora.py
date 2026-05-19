import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), "../../../results/bitacora/bitacora-mlp.xlsx")

results = [
    {
        "dataset": "X_scaled.csv", "arquitectura": "(128, 64, 32)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8540,
        "p0": 0.84, "r0": 0.82, "f0": 0.83,
        "p1": 0.86, "r1": 0.82, "f1": 0.84,
        "p2": 0.86, "r2": 0.92, "f2": 0.89,
        "cm": "[[273,34,25],[33,269,27],[18,9,312]]",
        "nota": "MEJOR CONFIGURACION"
    },
    {
        "dataset": "X_scaled.csv", "arquitectura": "(100, 50)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8490,
        "p0": 0.87, "r0": 0.77, "f0": 0.82,
        "p1": 0.84, "r1": 0.87, "f1": 0.85,
        "p2": 0.85, "r2": 0.90, "f2": 0.87,
        "cm": "[[257,38,37],[24,286,19],[16,17,306]]",
        "nota": ""
    },
    {
        "dataset": "X_scaled.csv", "arquitectura": "(200, 100)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8430,
        "p0": 0.86, "r0": 0.77, "f0": 0.82,
        "p1": 0.84, "r1": 0.83, "f1": 0.83,
        "p2": 0.83, "r2": 0.93, "f2": 0.88,
        "cm": "[[257,39,36],[29,272,28],[12,13,314]]",
        "nota": ""
    },
    {
        "dataset": "X.csv (raw)", "arquitectura": "(200, 100)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8510,
        "p0": 0.85, "r0": 0.78, "f0": 0.82,
        "p1": 0.82, "r1": 0.87, "f1": 0.85,
        "p2": 0.87, "r2": 0.90, "f2": 0.89,
        "cm": "[[259,49,24],[23,286,20],[21,12,306]]",
        "nota": "Mejor en dataset raw"
    },
    {
        "dataset": "X.csv (raw)", "arquitectura": "(100, 50)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8450,
        "p0": 0.86, "r0": 0.77, "f0": 0.81,
        "p1": 0.84, "r1": 0.84, "f1": 0.84,
        "p2": 0.84, "r2": 0.92, "f2": 0.88,
        "cm": "[[256,40,36],[28,277,24],[14,13,312]]",
        "nota": ""
    },
    {
        "dataset": "X.csv (raw)", "arquitectura": "(128, 64, 32)", "activacion": "ReLU", "optimizador": "Adam",
        "exactitud": 0.8420,
        "p0": 0.86, "r0": 0.77, "f0": 0.81,
        "p1": 0.84, "r1": 0.83, "f1": 0.83,
        "p2": 0.84, "r2": 0.92, "f2": 0.88,
        "cm": "[[255,45,32],[26,274,29],[17,9,313]]",
        "nota": ""
    },
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MLP Resultados"

# Styles
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
best_fill   = PatternFill("solid", fgColor="E2EFDA")
alt_fill    = PatternFill("solid", fgColor="F2F2F2")
center      = Alignment(horizontal="center", vertical="center")
thin        = Side(style="thin", color="BFBFBF")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Dataset", "Arquitectura", "Activacion", "Optimizador", "Exactitud",
    "P Clase 0", "R Clase 0", "F1 Clase 0",
    "P Clase 1", "R Clase 1", "F1 Clase 1",
    "P Clase 2", "R Clase 2", "F1 Clase 2",
    "Matriz de Confusion", "Nota"
]

ws.append(headers)
for col, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for i, r in enumerate(results, 2):
    row = [
        r["dataset"], r["arquitectura"], r["activacion"], r["optimizador"],
        r["exactitud"],
        r["p0"], r["r0"], r["f0"],
        r["p1"], r["r1"], r["f1"],
        r["p2"], r["r2"], r["f2"],
        r["cm"], r["nota"]
    ]
    ws.append(row)
    fill = best_fill if r["nota"] == "MEJOR CONFIGURACION" else (alt_fill if i % 2 == 0 else None)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=i, column=col)
        cell.border = border
        cell.alignment = center
        if fill:
            cell.fill = fill
        if col == 5:  # exactitud
            cell.number_format = "0.00%"

# Column widths
widths = [16, 18, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 36, 22]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 22

# Freeze header
ws.freeze_panes = "A2"

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Bitacora guardada: {OUT}")
